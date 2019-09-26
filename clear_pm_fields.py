import json
import requests
import warnings
from openpyxl import load_workbook


filepath = "C:\\Users\\cweaver\\Desktop\\Notes in the SIP tool\\Trouble_SIPs.xlsx"
wb=load_workbook(filepath)
sheet=wb.active
max_row=sheet.max_row

src_sip_id = []

for i in range(2, max_row + 1):
    # get particular cell value    
    if sheet.cell(row=i,column=5).value == False and sheet.cell(row=i, column=9).value == True:
        src_sip_id.append(sheet.cell(row=i, column=1).value)
print(src_sip_id)     
    


# warnings.simplefilter('ignore')

site = "https://avsip.ad.bl.uk"
# src_sip_id = [26429, 26373, 26265, 26264, 26263, 26262, 26261, 25863, 25862, 25861]
failed_sips = []


for sip in src_sip_id:
    try:
        # get information for the src
        src_url = "{0}/api/SIP/{1}".format(site, sip)
        #DEBUG print("getting information for sip", src_sip_id)
        src_req = requests.get(src_url, verify=False)
        if src_req.reason == 'Not Found':
            raise Exception(f'Process Metadata copy failed. Cannot find a SIP with ID number {sip}')
        src_req.encoding = src_req.apparent_encoding
        src_json = src_req.json()


        d = json.loads(src_json['ProcessMetadata'])

        for child in d['processMetadata'][0]['children']: 
            for device in child['devices']:
                print(device['notes'])
                device['notes'] = device['notes'].replace("Gawd the SIP tool's bollox", "")
                print(device['notes'])




        modified_process_metadata = json.dumps(d)
        # The following are required for the processMetadata API
        # https://avsip.ad.bl.uk/Help/Api/PATCH-api-SIP-processmetadata-sip_id-stepstate_id-user_id
        # 
        # get the user id
        src_user_id = src_json['UserId']
        
        # get the dest step state id for processmetadata
        for x in src_json['StepStates']:
            if x['StepTitle'] == "Process Metadata":
                src_step_id = x['StepStateId']
                break
        else:
            raise KeyError
        
        # make the patch request
        patch_url = "{0}/api/SIP/processmetadata/{1}/{2}/{3}/true".format(
            site,
            sip,
            src_step_id,
            src_user_id
        )
        print('patching', patch_url)
        r = requests.patch(
            patch_url,
            json=modified_process_metadata,
            verify=False
        )
    # # Once the Process Metadata has been copied from the JSON
    # # visit the page and step complete & continue
    # # this checks the process metadata

    # driver.get(f"{site}/Steps/Process/{dest_sip_id}/{dest_step_id}")
    # handle_logout("Process Metadata", "/Steps/Process/", dest_sip_id)

    # # Click 'OK' if there is a warning about Process Metadata vocabulary
    # time.sleep(2)
    # try:
    #     driver.find_element_by_xpath("//*[@class='modal-open']")
    #     driver.execute_script("arguments[0].click();", driver.find_element_by_xpath("//button[text()='OK']"))
    #     time.sleep(2)
    # except NoSuchElementException:
    #     pass
        

    # # Page is ready when this dropdown for custom SIP metadata is available
    # driver.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='process-metadata-form']/div[3]/button")))
    # driver.execute_script("arguments[0].click();", driver.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "step-complete-checkbox"))))
    # driver.execute_script("arguments[0].click();", driver.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='page-content-wrapper']/div[2]/div[2]/nav/button[3]"))))
    except Exception as e:
        failed_sips.append("SIP ID:" + str(sip) + ":  " + str(e))
        continue

if len(failed_sips) == 0:
    print("\n********************************************************************************")
    print("All SIPs completed.")
else:
    print("\n********************************************************************************")
    print("The following SIPs did not complete sucessfully:\n", )
    print(*failed_sips, sep="\n")




# print("\n********************************************************************************")
# print("\nProcess Metadata")
# print("\n")
# logger.info(f"Process Metadata complete for {site}/Steps/Process/{dest_sip_id}/{dest_step_id}")

# # Wait for the search box on the Search SAMI Recordings page
# # before starting next SIP otherwise you get an error 9the site isn't quite ready!)
# driver.wait.until(EC.visibility_of_element_located((By.ID, "searchBox")))
# time.sleep(1)
